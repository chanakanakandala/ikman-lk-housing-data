import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
import pandas as pd

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ASSETS_DIR = os.path.join(BASE_DIR, "..", "assets")
RAW_SCRAPE_DIR = os.path.join(BASE_DIR, "..", "raw_scrape")
CLEANED_SCRAPE_DIR = os.path.join(BASE_DIR, "..", "cleaned_scrape")
HISTORY_FILE = os.path.join(BASE_DIR, "..", "data/scrape_history.json")
LOCATIONS_FILE = os.path.join(ASSETS_DIR, "locations.json")


def load_history():
    """Load the scrape history from a JSON file."""
    if not os.path.exists(HISTORY_FILE):
        return []
    with open(HISTORY_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_history(history_data):
    """Overwrite the scrape_history.json file."""
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history_data, f, indent=4)


def add_history_record(record):
    """Append one record to scrape_history.json."""
    history = load_history()
    history.append(record)
    save_history(history)


def load_locations():
    """Read all available locations from assets/locations.json."""
    if not os.path.exists(LOCATIONS_FILE):
        return []
    with open(LOCATIONS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def ensure_excel_file():
    """
    Creates/returns today's raw Excel file (WB, sheet, path).
    """
    if not os.path.exists(RAW_SCRAPE_DIR):
        os.makedirs(RAW_SCRAPE_DIR)

    today_str = datetime.now().strftime("%Y-%m-%d")
    filename = os.path.join(RAW_SCRAPE_DIR, f"ikman_scrape_{today_str}.xlsx")

    if os.path.exists(filename):
        wb = load_workbook(filename)
        sheet = wb.active
        return wb, sheet, filename
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "ScrapeData"
        headers = [
            "Area Slug",
            "Location",
            "Title",
            "Description",
            "Details",
            "Price (numeric)",
            "Shop Name",
            "Slug",
            "URL",
            "Date"
        ]
        sheet.append(headers)
        wb.save(filename)
        return wb, sheet, filename


def append_to_excel(records):
    """
    Appends a list of records to today's raw Excel file.
    Each record is a list matching the columns above.
    """
    wb, sheet, filename = ensure_excel_file()
    for row_data in records:
        sheet.append(row_data)
    wb.save(filename)
    return filename


def read_excel_file(path):
    """Read an Excel file into a pandas DataFrame."""
    if not os.path.exists(path):
        return None
    return pd.read_excel(path)


def write_excel_file(df, path):
    """Write a pandas DataFrame to an Excel file."""
    # Ensure folder exists (for cleaned files, etc.)
    folder = os.path.dirname(path)
    if not os.path.exists(folder):
        os.makedirs(folder)

    df.to_excel(path, index=False)
