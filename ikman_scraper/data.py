import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ASSETS_DIR = os.path.join(BASE_DIR, "assets")
DATA_DIR = os.path.join(BASE_DIR, "data")
RAW_SCRAPE_DIR = os.path.join(BASE_DIR, "raw_scrape")
HISTORY_FILE = os.path.join(DATA_DIR, "scrape_history.json")
LOCATIONS_FILE = os.path.join(ASSETS_DIR, "locations.json")


def load_locations():
    """
    Load location data from assets/locations.json
    Expected format:
    [
      {"id": 10, "name": "Ampara", "slug": "ampara"},
      {"id": 1507, "name": "Athurugiriya", "slug": "athurugiriya"},
      ...
    ]
    """
    if not os.path.exists(LOCATIONS_FILE):
        return []
    with open(LOCATIONS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def ensure_excel_file():
    """
    Check if today's Excel file exists in 'raw_scrape' folder.
    If not, create with headers. Return (workbook, sheet, filename).
    """
    if not os.path.exists(RAW_SCRAPE_DIR):
        os.makedirs(RAW_SCRAPE_DIR)

    today_str = datetime.now().strftime("%Y-%m-%d")
    filename = os.path.join(RAW_SCRAPE_DIR, f"ikman_scrape_{today_str}.xlsx")

    if os.path.exists(filename):
        wb = load_workbook(filename)
        sheet = wb.active
        return wb, sheet, filename
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "ScrapeData"
        headers = [
            "Area Slug",
            "Date",
            "Location",
            "Title",
            "Description",
            "Details",
            "Price",
            "Agent Name",
            "URL",
            "Note"
        ]
        sheet.append(headers)
        wb.save(filename)
        return wb, sheet, filename


def append_to_excel(records):
    """Append a list of records to today's Excel file. Return the file path."""
    wb, sheet, filename = ensure_excel_file()
    for row_data in records:
        sheet.append(row_data)
    wb.save(filename)
    return filename


def clean_price(price_str):
    """
    Remove 'Rs', commas, spaces from price string, convert to int.
    e.g. "Rs 19,500,000" -> 19500000
    """
    if not price_str:
        return 0
    cleaned = price_str.replace("Rs", "").replace(",", "").strip()
    try:
        return int(cleaned)
    except ValueError:
        return 0


# -----------------------------
# HISTORY LOG FUNCTIONS
# -----------------------------
def load_history():
    """Load scrape history from scrape_history.json (list of dicts)."""
    if not os.path.exists(HISTORY_FILE):
        return []
    with open(HISTORY_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_history(history):
    """Overwrite scrape_history.json with the updated history."""
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history, f, indent=4)


def add_record_to_history(record):
    """
    Add one record (dict) to the existing scrape_history.json.
    record example:
    {
      "timestamp": "2025-01-05 12:34:56",
      "date": "2025-01-05",
      "locations_scraped": ["Ampara", "Athurugiriya"],
      "total_pages_scraped": 20,
      "total_ads_scraped": 123,
      "excel_file": ".../ikman_scrape_2025-01-05.xlsx"
    }
    """
    history = load_history()
    history.append(record)
    save_history(history)
